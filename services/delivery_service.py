from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import datetime, timezone

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from database.models import (
    ContactDelivery,
    DeliveredContact,
    Order,
    OrderStatus,
    User,
    UserStatus,
)
from services.user_service import order_remaining
from utils.phone import extract_phones


@dataclass
class DeliveryPreview:
    phones: list[str]
    duplicates: int
    invalid: int
    skipped_limit: int
    can_send: int


@dataclass
class DeliveryResult:
    delivery_id: int
    sent_count: int
    order_id: int
    user_db_id: int
    user_telegram_id: int
    received: int
    limit: int
    remaining: int
    order_completed: bool


def parse_contacts_file(content: bytes, filename: str) -> list[str]:
    name = filename.lower()
    if name.endswith(".xlsx"):
        return _parse_xlsx(content)
    if name.endswith(".csv"):
        return _parse_csv(content)
    return extract_phones(content.decode("utf-8", errors="ignore"))


def _parse_csv(content: bytes) -> list[str]:
    text = content.decode("utf-8-sig", errors="ignore")
    phones: list[str] = []
    reader = csv.reader(io.StringIO(text))
    for row in reader:
        phones.extend(extract_phones(" ".join(row)))
    return phones


def _parse_xlsx(content: bytes) -> list[str]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    phones: list[str] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            phones.extend(extract_phones(" ".join(str(c) for c in row if c is not None)))
    return phones


async def preview_delivery(
    session: AsyncSession,
    order_id: int,
    phones: list[str],
) -> DeliveryPreview:
    order = await session.get(Order, order_id)
    if not order:
        raise ValueError("Заказ не найден")

    remaining = order_remaining(order)
    seen: set[str] = set()
    valid: list[str] = []
    duplicates = 0

    existing = await session.execute(
        select(DeliveredContact.phone).where(DeliveredContact.order_id == order_id)
    )
    existing_phones = set(existing.scalars().all())

    for phone in phones:
        if phone in seen or phone in existing_phones:
            duplicates += 1
            continue
        seen.add(phone)
        valid.append(phone)

    can_send = min(len(valid), remaining)
    skipped = len(valid) - can_send
    invalid = len(phones) - len(valid) - duplicates

    return DeliveryPreview(
        phones=valid[:can_send],
        duplicates=duplicates,
        invalid=max(0, invalid),
        skipped_limit=skipped,
        can_send=can_send,
    )


async def commit_delivery(
    session: AsyncSession,
    order_id: int,
    phones: list[str],
    note: str | None = None,
) -> DeliveryResult:
    order = await session.get(Order, order_id)
    if not order:
        raise ValueError("Заказ не найден")

    user = await session.get(User, order.user_id)
    if not user:
        raise ValueError("Пользователь не найден")

    preview = await preview_delivery(session, order_id, phones)
    to_send = preview.phones
    if not to_send:
        raise ValueError("Нет контактов для отправки")

    delivery = ContactDelivery(
        user_id=user.id,
        order_id=order.id,
        count=len(to_send),
        note=note,
    )
    session.add(delivery)
    await session.flush()

    for phone in to_send:
        session.add(
            DeliveredContact(
                delivery_id=delivery.id,
                order_id=order.id,
                phone=phone,
            )
        )

    order.received += len(to_send)
    order.status = OrderStatus.IN_PROGRESS.value
    user.status = UserStatus.ACTIVE.value

    order_completed = order.received >= order.contact_limit
    if order_completed:
        order.status = OrderStatus.COMPLETED.value
        order.completed_at = datetime.now(timezone.utc)
        user.status = UserStatus.FINISHED.value

    await session.commit()
    await session.refresh(order)
    await session.refresh(user)

    return DeliveryResult(
        delivery_id=delivery.id,
        sent_count=len(to_send),
        order_id=order.id,
        user_db_id=user.id,
        user_telegram_id=user.telegram_id,
        received=order.received,
        limit=order.contact_limit,
        remaining=order_remaining(order),
        order_completed=order_completed,
    )
